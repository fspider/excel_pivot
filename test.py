from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils import get_column_letter


prefix = 'UBILLS TEST'
outfile = prefix + '.xlsx'


wb = load_workbook(outfile)
print (wb.sheetnames)
sheet_abs = wb['SUMM-ABS Status']

row_count = sheet_abs.max_row
column_count = sheet_abs.max_column
print(row_count, column_count)
sheet_abs.insert_rows(1, 2)

# for items in sorted(sheet_abs.merged_cell_ranges):
#   print(items)
#   sheet_abs.unmerge_cells(str(items))
#   sheet.unmerge_cells(str(items))

# sheet_abs.merge_cells(start_row=2, start_column=1, end_row=2, end_column=column_count)
# sheet_abs.delete_rows(1)
# my_cell = sheet_abs.cell(1, 1)
# my_cell.value = "My Cell"
wb.save('test_output.xlsx')
